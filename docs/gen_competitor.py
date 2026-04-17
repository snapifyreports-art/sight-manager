import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

green_fill = PatternFill('solid', fgColor='C6EFCE')
yellow_fill = PatternFill('solid', fgColor='FFEB9C')
red_fill = PatternFill('solid', fgColor='FFC7CE')
header_fill = PatternFill('solid', fgColor='1F4E79')
section_fill = PatternFill('solid', fgColor='D6E4F0')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
section_font = Font(name='Arial', bold=True, color='1F4E79', size=10)
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

def style_header(ws, row, cols):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_section(ws, row, cols):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border

def style_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if value == 'Y':
        cell.fill = green_fill
        cell.value = 'YES'
        cell.alignment = Alignment(horizontal='center', vertical='center')
    elif value == 'P':
        cell.fill = yellow_fill
        cell.value = 'PARTIAL'
        cell.alignment = Alignment(horizontal='center', vertical='center')
    elif value == 'N':
        cell.fill = red_fill
        cell.value = 'NO'
        cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        cell.value = value

# === SHEET 1: Feature Comparison ===
ws1 = wb.active
ws1.title = 'Feature Comparison'
headers = ['Category', 'Feature', 'Sight Manager', 'Procore', 'Fieldwire', 'Buildertrend', 'Archdesk', 'Asta Powerproject']
for i, h in enumerate(headers, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, 8)

features = [
    ('PROGRAMME & SCHEDULING', None, None, None, None, None, None, None),
    (None, 'Gantt chart', 'Y', 'Y', 'P', 'Y', 'Y', 'Y'),
    (None, 'Original vs Current overlay', 'Y', 'N', 'N', 'N', 'N', 'Y'),
    (None, 'Cascade/auto-shift dates', 'Y', 'P', 'N', 'P', 'P', 'Y'),
    (None, 'Working days (Mon-Fri)', 'Y', 'N', 'N', 'N', 'N', 'Y'),
    (None, 'Critical path analysis', 'Y', 'P', 'N', 'N', 'N', 'Y'),
    (None, 'Programme templates', 'Y', 'Y', 'Y', 'Y', 'Y', 'Y'),
    (None, 'Multi-plot management', 'Y', 'P', 'N', 'P', 'P', 'P'),
    (None, 'Weather impact on schedule', 'Y', 'N', 'N', 'N', 'N', 'N'),
    (None, 'Rained off day tracking', 'Y', 'N', 'N', 'N', 'N', 'N'),
    ('DAILY OPERATIONS', None, None, None, None, None, None, None),
    (None, 'Daily brief / morning dashboard', 'Y', 'P', 'N', 'P', 'N', 'N'),
    (None, 'Site walkthrough mode', 'Y', 'N', 'P', 'N', 'N', 'N'),
    (None, 'Quick start/complete/sign-off', 'Y', 'P', 'P', 'P', 'P', 'N'),
    (None, 'Readiness checklist before start', 'Y', 'N', 'N', 'N', 'N', 'N'),
    (None, 'Post-completion workflow', 'Y', 'N', 'N', 'N', 'N', 'N'),
    (None, 'Bulk actions', 'Y', 'Y', 'P', 'P', 'N', 'N'),
    (None, 'Real-time weather integration', 'Y', 'N', 'N', 'P', 'N', 'N'),
    (None, 'Mobile-first design', 'Y', 'Y', 'Y', 'Y', 'P', 'N'),
    ('ORDER & MATERIAL MANAGEMENT', None, None, None, None, None, None, None),
    (None, 'Full order lifecycle', 'Y', 'P', 'N', 'Y', 'Y', 'N'),
    (None, 'Email supplier from app', 'Y', 'P', 'N', 'Y', 'P', 'N'),
    (None, 'Order templates from jobs', 'Y', 'N', 'N', 'P', 'P', 'N'),
    (None, 'Delivery date tracking', 'Y', 'P', 'N', 'Y', 'Y', 'N'),
    (None, 'Overdue delivery alerts', 'Y', 'N', 'N', 'P', 'P', 'N'),
    (None, 'Inline order editing', 'Y', 'Y', 'N', 'Y', 'Y', 'N'),
    (None, 'Order cost tracking', 'Y', 'Y', 'N', 'Y', 'Y', 'N'),
    (None, 'PO numbering', 'N', 'Y', 'N', 'Y', 'Y', 'N'),
    (None, 'Partial delivery recording', 'N', 'P', 'N', 'P', 'P', 'N'),
    ('SNAG MANAGEMENT', None, None, None, None, None, None, None),
    (None, 'Raise snags with photos', 'Y', 'Y', 'Y', 'Y', 'P', 'N'),
    (None, 'Before/after photo tagging', 'Y', 'P', 'P', 'N', 'N', 'N'),
    (None, 'Assign snag to contractor', 'Y', 'Y', 'Y', 'P', 'P', 'N'),
    (None, 'Snag resolution workflow', 'Y', 'Y', 'Y', 'P', 'P', 'N'),
    (None, 'Snag ageing report', 'Y', 'P', 'N', 'N', 'N', 'N'),
    (None, 'Filters (contractor/plot/priority)', 'Y', 'Y', 'Y', 'P', 'P', 'N'),
    (None, 'Auto-assign contractor from job', 'Y', 'N', 'N', 'N', 'N', 'N'),
    ('CONTRACTOR MANAGEMENT', None, None, None, None, None, None, None),
    (None, 'Contractor comms hub', 'Y', 'P', 'N', 'P', 'P', 'N'),
    (None, 'Shareable day sheets/links', 'Y', 'N', 'N', 'N', 'N', 'N'),
    (None, 'Contractor confirmation workflow', 'Y', 'N', 'N', 'N', 'N', 'N'),
    (None, 'Call/email from card', 'Y', 'Y', 'P', 'Y', 'P', 'N'),
    (None, 'Contractor portal (read-only)', 'Y', 'Y', 'N', 'Y', 'N', 'N'),
    ('REPORTING & ANALYTICS', None, None, None, None, None, None, None),
    (None, 'Budget vs actual', 'Y', 'Y', 'N', 'Y', 'Y', 'P'),
    (None, 'Cash flow forecast', 'Y', 'Y', 'N', 'P', 'Y', 'N'),
    (None, 'Delay/blame report', 'Y', 'N', 'N', 'N', 'N', 'P'),
    (None, 'Weekly progress report', 'Y', 'Y', 'P', 'Y', 'P', 'Y'),
    (None, 'Site heatmap', 'Y', 'N', 'N', 'N', 'N', 'N'),
    (None, 'Export PDF/Excel', 'Y', 'Y', 'Y', 'Y', 'Y', 'Y'),
    ('COLLABORATION', None, None, None, None, None, None, None),
    (None, 'Multi-user with roles', 'Y', 'Y', 'Y', 'Y', 'Y', 'Y'),
    (None, 'Event/activity log', 'Y', 'Y', 'P', 'Y', 'Y', 'P'),
    (None, 'Push notifications', 'Y', 'Y', 'Y', 'Y', 'P', 'N'),
    (None, 'Shareable read-only links', 'Y', 'P', 'N', 'P', 'N', 'N'),
    (None, 'Site diary / daily log', 'N', 'Y', 'Y', 'Y', 'P', 'N'),
    ('INTEGRATIONS', None, None, None, None, None, None, None),
    (None, 'Weather API', 'Y', 'N', 'N', 'P', 'N', 'N'),
    (None, 'Excel import/export', 'Y', 'Y', 'P', 'Y', 'Y', 'Y'),
    (None, 'Accounting integration', 'N', 'Y', 'N', 'Y', 'Y', 'N'),
    (None, 'Document management', 'P', 'Y', 'Y', 'Y', 'Y', 'P'),
    (None, 'H&S module', 'N', 'Y', 'Y', 'P', 'P', 'N'),
    (None, 'BIM integration', 'N', 'Y', 'P', 'N', 'N', 'Y'),
]

row = 2
for f in features:
    if f[1] is None:
        ws1.cell(row=row, column=1, value=f[0])
        style_section(ws1, row, 8)
    else:
        ws1.cell(row=row, column=1, value='').font = normal_font
        ws1.cell(row=row, column=1).border = thin_border
        ws1.cell(row=row, column=2, value=f[1]).font = normal_font
        ws1.cell(row=row, column=2).border = thin_border
        for col_idx in range(3, 9):
            style_cell(ws1, row, col_idx, f[col_idx-1])
    row += 1

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 35
for c in range(3, 9):
    ws1.column_dimensions[get_column_letter(c)].width = 16
ws1.freeze_panes = 'C2'

# === SHEET 2: Pricing ===
ws2 = wb.create_sheet('Pricing Comparison')
p_headers = ['Company', 'Starting Price', 'Mid-Tier', 'Enterprise', 'Pricing Model', 'Free Trial', 'UK Focused', 'Best For']
for i, h in enumerate(p_headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, 8)

pricing = [
    ('Sight Manager', '\u00a3349/site/mo', '\u00a3499/site/mo', 'Custom', 'Per site + included users', '30-day free pilot', 'YES', 'UK housebuilders 1-50 sites'),
    ('Procore', '~$375/mo (ACV)', '$10-25k/year', '$25k+/year', 'Annual construction volume', 'Demo only', 'No (US)', 'Large GCs, commercial'),
    ('Fieldwire', '$54/user/mo', '$74/user/mo', '$94/user/mo', 'Per user/month', 'Free tier', 'No (US)', 'Field teams, tasks'),
    ('Buildertrend', '$199-499/mo', '$499-799/mo', 'Custom', 'Per company/month', '14-day trial', 'No (US)', 'US home builders'),
    ('Archdesk', '$790/mo', 'Custom', 'Custom', 'Per company', 'Demo only', 'Yes (UK/EU)', 'UK contractors'),
    ('Asta Powerproject', '$29/user/mo', '$99-199/user/mo', '$1,449 perpetual', 'Per user or licence', '14-day trial', 'Yes (UK)', 'Planning teams'),
]

for r, p in enumerate(pricing, 2):
    for c, v in enumerate(p, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = bold_font if c == 1 else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

for c in range(1, 9):
    ws2.column_dimensions[get_column_letter(c)].width = 22
ws2.freeze_panes = 'B2'

# === SHEET 3: Strengths & Weaknesses ===
ws3 = wb.create_sheet('Strengths & Weaknesses')
sw_headers = ['Company', 'Strength 1', 'Strength 2', 'Strength 3', 'Weakness 1', 'Weakness 2', 'Weakness 3', 'Best For', 'UK Fit (1-10)']
for i, h in enumerate(sw_headers, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, 9)

sw_data = [
    ('Sight Manager', 'Purpose-built for UK housebuilders', 'Full order lifecycle + programme cascade', 'Daily Brief + Walkthrough for site ops', 'No accounting integration', 'No H&S module', 'No document management', 'UK housebuilders 1-50 sites', '9'),
    ('Procore', 'Enterprise unlimited users', 'Comprehensive module ecosystem', 'Strong document management', 'Very expensive for small builders', 'US-focused', 'No residential workflows', 'Large commercial GCs', '3'),
    ('Fieldwire', 'Excellent mobile experience', 'Strong plan/drawing management', 'Good task management', 'No programme/Gantt', 'No order management', 'No UK features', 'Field teams', '2'),
    ('Buildertrend', 'Strong residential focus', 'Client portal + selections', 'Integrated estimating', 'US-centric', 'No cascade', 'Expensive', 'US custom builders', '4'),
    ('Archdesk', 'UK-based', 'Good job costing', 'Full lifecycle', 'Expensive starting point', 'Not residential focused', 'Less intuitive UX', 'UK commercial contractors', '5'),
    ('Asta Powerproject', 'Best scheduling/planning', 'Critical path + earned value', '4D BIM', 'Planning only', 'No site operations', 'Steep learning curve', 'Planning departments', '4'),
]

for r, d in enumerate(sw_data, 2):
    for c, v in enumerate(d, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.font = bold_font if c == 1 else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

for c in range(1, 10):
    ws3.column_dimensions[get_column_letter(c)].width = 25
ws3.freeze_panes = 'B2'

# === SHEET 4: They Have We Don't ===
ws4 = wb.create_sheet('They Have We Dont')
th_headers = ['Feature', 'Which Competitors', 'Priority', 'Effort', 'Business Impact', 'Notes']
for i, h in enumerate(th_headers, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, 6)

they_have = [
    ('PO numbering', 'Procore, Buildertrend, Archdesk', 'P1', 'Low', 'High', 'Auto-generate SITE-PO-001'),
    ('Partial delivery recording', 'Procore, Buildertrend, Archdesk', 'P1', 'Medium', 'High', 'Qty received vs ordered, condition notes'),
    ('Accounting integration', 'Procore, Buildertrend, Archdesk', 'P2', 'High', 'High', 'Xero/Sage/QuickBooks API'),
    ('Document management', 'Procore, Fieldwire, Buildertrend, Archdesk', 'P2', 'Medium', 'Medium', 'Start with contractor certificates'),
    ('Site diary / daily log', 'Procore, Fieldwire, Buildertrend', 'P2', 'Medium', 'High', 'Auto-populate + manual entry'),
    ('H&S module', 'Procore, Fieldwire', 'P3', 'High', 'Medium', 'Inspection checklists + incidents'),
    ('BIM integration', 'Procore, Asta', 'P3', 'Very High', 'Low for residential', 'Not priority'),
    ('Client/buyer portal', 'Buildertrend', 'P3', 'Medium', 'Medium', 'Read-only plot progress'),
    ('Estimating / takeoff', 'Procore, Buildertrend, Archdesk', 'P3', 'Very High', 'Medium', 'Consider integration'),
    ('AI agents', 'Procore', 'P3', 'High', 'Future', 'Auto-scheduling, predictions'),
]

for r, d in enumerate(they_have, 2):
    for c, v in enumerate(d, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if c == 3:
            if v == 'P1': cell.fill = PatternFill('solid', fgColor='FF6B6B')
            elif v == 'P2': cell.fill = PatternFill('solid', fgColor='FFD93D')
            elif v == 'P3': cell.fill = PatternFill('solid', fgColor='6BCB77')

for c in range(1, 7):
    ws4.column_dimensions[get_column_letter(c)].width = 30
ws4.freeze_panes = 'A2'

# === SHEET 5: Our Unique Advantages ===
ws5 = wb.create_sheet('Our Unique Advantages')
ua_headers = ['Feature', 'Why Unique', 'Why It Matters', 'Competitor Gap']
for i, h in enumerate(ua_headers, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, 4)

advantages = [
    ('Daily Brief', 'No competitor combines weather + jobs + materials + snags + contractor status in one actionable morning view', 'Site managers check 4-5 tools. We give them ONE screen at 7am.', 'Procore dashboards are project-level. Nobody has readiness checklists.'),
    ('Site Walkthrough', 'Plot-by-plot mobile inspection with integrated snag/note/photo/job actions', 'Walking a site is how managers check progress. No tool supports this.', 'Fieldwire has tasks but no walkthrough. Nobody has plot-by-plot navigation.'),
    ('Programme cascade', 'Auto-shift all jobs + orders using Mon-Fri working days', 'Programme changes daily. Manual re-scheduling takes hours.', 'Asta has cascade but 10x price. Procore scheduling is basic.'),
    ('Order + Programme integration', 'Full lifecycle linked to job start/completion with auto-progression', 'Material orders = #1 delay cause. Linking to programme prevents misses.', 'Buildertrend has POs not linked to cascade. No competitor auto-progresses.'),
    ('Contractor shareable links', 'Free read-only links, no login, live data', 'Contractors hate apps. A browser link removes all friction.', 'Procore requires paid accounts. Buildertrend requires sign-up.'),
    ('Rained off tracking', 'Mark rained off + auto programme delay + weather alerts', 'UK weather = 10-15% of delays. No other tool tracks this.', 'No competitor has this.'),
    ('Post-completion workflow', 'Smart dialog: start next, Monday, push, defer', 'Keeps programme moving without manual re-planning.', 'No competitor has this.'),
    ('Readiness checklist', 'Checks contractor, orders, materials, predecessor before start', 'Prevents starting without prerequisites = #1 idle crew cause.', 'No competitor checks readiness.'),
    ('Delay blame analysis', 'Auto-split delays by contractor, weather, supplier', 'Essential for negotiations and recovery planning.', 'No competitor has automated blame.'),
    ('UK housebuilder focus', 'Plot-based, UK weather, UK terms, UK workflows', 'Every competitor is US or generic. None understand UK housebuilding.', 'Zero competitors purpose-built for UK residential.'),
]

for r, d in enumerate(advantages, 2):
    for c, v in enumerate(d, 1):
        cell = ws5.cell(row=r, column=c, value=v)
        cell.font = bold_font if c == 1 else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

for c in range(1, 5):
    ws5.column_dimensions[get_column_letter(c)].width = 40
ws5.freeze_panes = 'B2'

output = r'C:\Users\keith\OneDrive\Desktop\sight-manager\docs\Competitor-Analysis.xlsx'
wb.save(output)
print(f'Saved: {output}')
