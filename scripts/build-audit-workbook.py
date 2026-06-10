# Builds docs/Sight-Manager-Audit-Decisions.xlsx — plain-English tick sheet.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

F = "Arial"
HEAD_FILL = PatternFill("solid", start_color="1F3864")
HEAD_FONT = Font(name=F, bold=True, color="FFFFFF", size=11)
ADVICE_FILL = PatternFill("solid", start_color="E2EFDA")
ANSWER_FILL = PatternFill("solid", start_color="FFF2CC")
DONE_FILL = PatternFill("solid", start_color="D9EAD3")
BODY = Font(name=F, size=10)
BOLD = Font(name=F, size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

wb = Workbook()

def style_sheet(ws, widths, n_rows):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=len(widths)):
        for c in row:
            c.border = THIN
            if c.row == 1:
                c.fill = HEAD_FILL; c.font = HEAD_FONT
                c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            else:
                if c.font is None or not c.font.bold: c.font = BODY
                c.alignment = WRAP
    ws.freeze_panes = "A2"

# ── READ ME ─────────────────────────────────────────────────────────────
rm = wb.active; rm.title = "READ ME"
rm_rows = [
    ["Sight Manager — audit decisions (10 June 2026)"],
    [""],
    ["What this is:"],
    ["I swept the whole app for the four problem types you spotted (dead buttons, codes shown instead of names,"],
    ["cut-down forms, clunky flows). Everything that was clearly broken is ALREADY FIXED (green tab)."],
    ["What's left are choices only you can make, plus optional improvements."],
    [""],
    ["How to use it:"],
    ["1. Open the DECISIONS tab. Read each row — plain English, no jargon."],
    ["2. The green column is what I'd do. If you agree, pick \"Do your advice\" in the orange column (it's a dropdown)."],
    ["   If you prefer another option, pick it. Add anything else in Your notes."],
    ["3. Do the same on QUICK WINS — just Yes / No / Ask me first."],
    ["4. Save the file and send it back (or just tell me the IDs, e.g. \"D1 advice, D4 yes, W1-W5 yes\")."],
    [""],
    ["The GUARDRAILS tab answers your question about stopping these bugs coming back."],
]
for r in rm_rows: rm.append(r)
rm["A1"].font = Font(name=F, bold=True, size=14)
for label in ("A3", "A8"): rm[label].font = Font(name=F, bold=True, size=11)
rm.column_dimensions["A"].width = 110
for row in rm.iter_rows(min_row=1, max_row=len(rm_rows), max_col=1):
    for c in row:
        if c.font is None or c.font.size is None: c.font = BODY

# ── DECISIONS ───────────────────────────────────────────────────────────
ds = wb.create_sheet("DECISIONS")
ds.append(["ID", "What's going on (plain English)", "Why it matters",
           "MY ADVICE — what I'd do", "Other choices", "Your answer", "Your notes"])

decisions = [
    ("D1",
     "When a job can't be finished because an inspection hasn't passed, you can type a reason to push it through anyway. But after you do, the notes and photos you'd typed for the sign-off get thrown away and you have to start again.",
     "This is the flow you'll use every day on site. Losing typed notes is infuriating and looks broken in front of the site manager.",
     "One tap: after you give the override reason, the app finishes the job AND completes your sign-off with your notes and photos automatically.",
     "OR: come back to the sign-off screen with a message saying 'job finished — tap Sign Off again'. (Two taps, but more explicit.)"),
    ("D2",
     "Every 'add photo' button in the app forces the phone camera open. You can never pick a photo you took earlier, or one a contractor sent you on WhatsApp.",
     "Half the photos a manager wants to attach already exist on the phone.",
     "Change every photo button to show the normal phone choice: Camera or Photo Library. Camera is still one tap.",
     "OR: do that everywhere EXCEPT the Site Walkthrough, which stays camera-first for speed. OR: leave camera-only."),
    ("D3",
     "After signing off a job on the Daily Brief, a small message pops up for 10 seconds offering 'review next steps' (start the next job, pull work forward). Miss it and it's gone.",
     "That prompt is where good scheduling decisions happen — it shouldn't vanish.",
     "Put a little 'Review next steps' button on the signed-off job row that stays until you've dealt with it (keep the pop-up too).",
     "OR: open the next-steps screen automatically every time. OR: just make the pop-up last 20 seconds."),
    ("D4",
     "Inspections you add by hand from a plot are second-class: they can't block the job like template ones can, no booking reminder, no inspector, no notes.",
     "A one-off Building Control visit added on the fly is just as real as a planned one.",
     "Give the manual 'add inspection' form everything the template form has — including the hard-block option. (Small server change needed.)",
     "OR: add notes/inspector/reminders but keep hard-blocking template-only. OR: leave one-offs as simple reminders."),
    ("D5",
     "When an inspection fails and creates an NCR (a formal quality record), there's nowhere to write the root cause or how it was put right — and those boxes stay blank in the handover pack forever.",
     "The QA paperwork going to a buyer or warranty provider shouldn't have permanently blank boxes.",
     "When raising an NCR from an inspection, show two optional boxes — Root cause and How it was fixed. They already save fine on the normal NCR screen.",
     "OR: leave it; fill them in later once an NCR edit screen exists (there isn't one yet)."),
    ("D6",
     "NCRs raised by hand can't be linked to a plot, job, or contractor — so they show '—' in the register and never count against the contractor's scorecard.",
     "Without the link, repeat offenders don't show up in their numbers.",
     "Add three optional pickers (Plot / Job / Contractor) to the Raise NCR form. The plumbing already exists.",
     "OR: leave manual NCRs site-level only."),
    ("D7",
     "The 'Notify contractor' button on an inspection quietly does nothing useful if that contractor has no email saved — and it can only message one contractor, with no attachments.",
     "You think the trade has been told about the inspection. They haven't.",
     "Right now: mark contractors with no email in the picker so you can see it before sending. Later: reuse the full Toolbox Talk form (multiple people, attachments).",
     "OR: do the full reuse now. OR: leave as is."),
    ("D8",
     "The 'Book' button on an inspection assumes the inspector is coming on the planned date. If they offer a different day there's no way to record it.",
     "The contractor share page and 'who's on site today' show the wrong visit day.",
     "Tapping Book shows one date box, pre-filled with the planned date — press Enter to keep it (as quick as today) or change it.",
     "OR: keep one-tap Book and let you edit the date afterwards on the row. OR: leave as is."),
    ("D9",
     "Some labels look like shouting computer codes ('AWAITING CORRECTION'). I've fixed all the places that showed raw codes; this is about the remaining CAPITALS styling on little status chips.",
     "Purely cosmetic consistency.",
     "Keep the capital-letters LOOK on the chips (it's a style), but make the underlying text proper words everywhere so tooltips, PDFs and emails always read nicely.",
     "OR: change every chip to normal 'Title Case' text. OR: leave the chips alone entirely."),
    ("D10",
     "A few old entries in the Site Log were written with codes/IDs in them before I fixed the cause. New entries are clean; the old ones stay as they were.",
     "History only — nothing new is affected.",
     "Leave the old entries alone (fix-forward).",
     "OR: run a one-off clean-up of old entries. OR: tidy them up only when displayed."),
    ("D11",
     "The 'Open in Email App' button for sending a contractor their link still works even when the contractor has no email saved — it opens an empty email going to nobody.",
     "Easy to think you've sent something you haven't.",
     "Grey the button out when there's no email saved (a hint already says to add one).",
     "OR: leave it — the hint below is enough."),
    ("D12",
     "Two chunks of dead code: an old 'Jobs' page that isn't reachable and whose Create button never worked, and a 'shared photo button' component nothing actually uses.",
     "Dead code confuses future work and can get accidentally wired back in, broken.",
     "Delete both.",
     "OR: delete the jobs page but keep + properly adopt the shared photo button (bigger job). OR: keep both."),
]
for d in decisions:
    ds.append([d[0], d[1], d[2], d[3], d[4], "", ""])

n = len(decisions) + 1
style_sheet(ds, [6, 46, 30, 40, 36, 16, 24], n)
for r in range(2, n + 1):
    ds.cell(row=r, column=1).font = BOLD
    ds.cell(row=r, column=4).fill = ADVICE_FILL
    ds.cell(row=r, column=6).fill = ANSWER_FILL
dv = DataValidation(type="list", formula1='"Do your advice,Other option,Skip for now,Talk to me first"', allow_blank=True)
ds.add_data_validation(dv)
dv.add(f"F2:F{n}")

# ── QUICK WINS ──────────────────────────────────────────────────────────
qw = wb.create_sheet("QUICK WINS")
qw.append(["ID", "What it would do (plain English)", "Why it helps on site", "Size", "Build it?", "Your notes"])
wins = [
    ("W1", "When you hit 'Add Snag' from a plot you already chose, skip the pointless 'pick a plot' step.", "One less tap on the busiest action.", "Small"),
    ("W2", "On the Site Closure checklist, each warning ('3 open snags') gets a 'Fix →' link straight to the right tab.", "No hunting for where to fix things.", "Small"),
    ("W3", "Inspection rows on a plot jump straight to that exact inspection, not the big global list.", "Less scrolling and searching.", "Small"),
    ("W4", "The 'Confirm pass' button stays greyed until you've picked a certificate (instead of an error after you've filled everything in).", "No wasted effort at sign-off.", "Small"),
    ("W5", "The certificate picker lists actual certificates first (not every document on the plot).", "Right file in one tap.", "Small"),
    ("W6", "Walkthrough snag photos automatically tagged 'before', plus optional assign-to and notes.", "Better evidence trail with zero extra effort.", "Small"),
    ("W7", "Document upload gets an optional type dropdown (Certificate / Drawing / RAMS).", "Files land in the right handover folder by themselves.", "Small"),
    ("W8", "The 'job' picker when creating an order becomes type-to-search (today it's one giant list of every job on every site).", "Usable on a phone with big portfolios.", "Medium"),
    ("W9", "The single-order form gets the same itemised lines (qty/cost) the bulk wizard has.", "Costs tracked the same whichever way you order.", "Medium"),
    ("W10", "A 'Send from Sight Manager' option for contractor links (today it only opens your own email app).", "Sent links get recorded properly in the app.", "Medium"),
    ("W11", "DONE — every deploy now automatically checks that no button/link points at a page that doesn't exist. (It already caught and fixed two.)", "Stops the '404 button' bug class permanently.", "Done"),
    ("W12", "DONE — a shared 'plain words' translator for all status codes + an automatic warning when code tries to show raw codes again.", "Stops the 'codes instead of names' bug class.", "Done"),
    ("W13", "Finish moving the last few screens onto the shared label translator (the cosmetic CAPITALS chips from D9).", "One source of wording everywhere.", "Small"),
]
for w in wins:
    qw.append([w[0], w[1], w[2], w[3], "Done" if w[3] == "Done" else "", ""])
n = len(wins) + 1
style_sheet(qw, [6, 56, 34, 10, 14, 24], n)
for r in range(2, n + 1):
    qw.cell(row=r, column=1).font = BOLD
    if qw.cell(row=r, column=4).value == "Done":
        for c in range(1, 7): qw.cell(row=r, column=c).fill = DONE_FILL
    else:
        qw.cell(row=r, column=5).fill = ANSWER_FILL
dv2 = DataValidation(type="list", formula1='"Yes,No,Ask me first"', allow_blank=True)
qw.add_data_validation(dv2)
dv2.add(f"E2:E{n}")

# ── ALREADY FIXED ───────────────────────────────────────────────────────
af = wb.create_sheet("ALREADY FIXED")
af.append(["What was wrong (plain English)", "Where you'd have hit it", "Status"])
fixed = [
    ("Every 'share link' button in the app (contractor View/Send Link, customer links, calendar links, cabin TV links) was crashing with an error — a server setting name mismatch.", "Contractor Comms — your 'View buttons not working'", "Fixed + tested live"),
    ("The delete button on site Documents did nothing — the 'are you sure?' box never appeared.", "Site Admin → Documents — your report", "Fixed + tested live"),
    ("Snags raised during an inspection couldn't have photos/location/assignee. Now a 'Snag + photos' button opens the full snag form, auto-linked to the inspection.", "Inspections → Pass/Fail — your report", "Fixed"),
    ("Plot names on 'Who's on site today' linked to a page that doesn't exist (error page).", "On Site Today", "Fixed"),
    ("Logged-out users opening a Walkthrough link were sent to a page that doesn't exist instead of the login screen.", "Site Walkthrough", "Fixed (caught by the new automatic link check)"),
    ("27 action buttons were slightly too small to tap reliably on a phone (a typo in their styling).", "Daily Brief + plot to-do lists", "Fixed"),
    ("The Events Log showed raw computer codes (e.g. INSPECTION_FAILED) for 13 kinds of event, and you couldn't filter to them.", "Events Log", "Fixed"),
    ("Plot IDs (long random codes) were being written permanently into the Site Log when a plot had no number.", "Site Log / Site Story — your 'IDs not names'", "Fixed"),
    ("18 places showed codes instead of words — handover PDFs (13 spots), story timeline, analytics, dashboard, customer page, calendar feed, weekly email, search results.", "All over — your 'IDs not names'", "Fixed"),
    ("Uploading a 10–50MB drawing was wrongly rejected with a misleading '10MB limit' message.", "Documents upload", "Fixed"),
    ("'Notify contractor' and certificate upload failed silently — you never knew it hadn't worked.", "Inspections", "Fixed (now shows the reason)"),
    ("The sign-off photo button said 'or choose from gallery' but forced the camera.", "Daily Brief sign-off", "Fixed (this one screen; D2 decides app-wide)"),
    ("Retrying after a partly-failed site creation quietly created a DUPLICATE site and duplicate plots.", "New site wizard", "Fixed (retry now reuses the site and only redoes the failed plots)"),
    ("Signing off after an inspection override dead-ended silently and lost your notes (the 'button does nothing' half).", "Daily Brief sign-off", "Fixed (D1 decides the one-tap flow)"),
]
for f_ in fixed: af.append(list(f_))
n = len(fixed) + 1
style_sheet(af, [70, 34, 26], n)
for r in range(2, n + 1):
    af.cell(row=r, column=3).fill = DONE_FILL

# ── GUARDRAILS ──────────────────────────────────────────────────────────
gr = wb.create_sheet("GUARDRAILS")
gr.append(["Why bugs like this happened", "The fence now in place", "Status"])
fences = [
    ("Buttons/links pointing at pages that don't exist (your dead View buttons, the 404 plot links).",
     "Every deploy now runs an automatic check of every internal link against the real list of pages. If anyone adds a dead link, the deploy FAILS before it reaches you. It already caught a second bug (the login redirect) the moment it was switched on.",
     "LIVE — runs on every deploy"),
    ("Computer codes leaking into screens, PDFs and emails (your 'IDs not names').",
     "One shared 'plain words' translator file now exists that every screen, PDF and email uses. A code-style warning fires the moment a developer tries the lazy underscore-swap trick instead of using it.",
     "LIVE — warning now; becomes a hard block after D9 tidy-up"),
    ("Confirm boxes that never appeared (your dead Documents delete button).",
     "All 16 screens using confirm boxes were checked one by one for the same mistake (only one had it). A deeper fix — making the confirm box impossible to forget — is recommended below.",
     "Swept clean; deeper fix recommended"),
    ("Cut-down copies of forms drifting from the real one (your snag-without-photos).",
     "Principle adopted: one true form per thing, reused everywhere with presets (done for snags-from-inspections). The remaining clones are listed as decisions D4-D7 so they converge too.",
     "In progress via D4-D7"),
    ("Server settings with two possible names (the crash behind your dead View buttons).",
     "The code now accepts both names, so this exact crash can't recur. Recommended extra: a start-up check that lists any missing settings clearly instead of crashing mid-click.",
     "Fixed; start-up check recommended"),
    ("Big changes shipped fast (the inspections build) introduce small regressions.",
     "Working pattern now: every batch is independently reviewed by a second pass before deploy (this caught 13 issues before you saw them), then click-tested on the live site.",
     "Standing practice"),
]
for f_ in fences: gr.append(list(f_))
n = len(fences) + 1
style_sheet(gr, [44, 70, 26], n)
for r in range(2, n + 1):
    if "LIVE" in str(gr.cell(row=r, column=3).value): gr.cell(row=r, column=3).fill = DONE_FILL

out = "docs/Sight-Manager-Audit-Decisions.xlsx"
wb.save(out)
print("saved", out)
